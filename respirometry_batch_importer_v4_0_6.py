#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Respirometry batch importer - v4.0.6
# Nowość:
# - JSON cache: processed_files.json (domyślnie w --input-dir), aby pomijać pliki bez zmian (po nazwie, rozmiarze i mtime)
# - Wszystkie poprawki v4.0.5 (brak % inhibicji dla referencji, problems_log, logi meta, aliasy, S/T, rozszerzanie tabel)
import argparse, sys, re, shutil, zipfile, xml.etree.ElementTree as ET, math, json
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import numpy as np, pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

RAW_SHEET_CANDIDATES = ["dane surowe","dane surowe ","dane  surowe","dane surowe  ","Dane surowe"]
DATE_EXCEL_FMT = "DD.MM.YYYY"
VERSION = "v4.0.6"

SEGMENT_ALIAS = {
    "Data": ["Data"],
    "Dawka Cr [mg]": ["Dawka Cr [mg]"],
    "Respirometr/wersja": ["Respirometr/wersja"],
    "sm (średnia) g sm/L": ["sm (średnia) g sm/L", "Sucha masa (średnia) [g sm/L]", "Sucha masa [g sm/L]"],
    "t_start [s]": ["t_start [s]"],
    "t_end [s]": ["t_end [s]"],
    "Δt [s]": ["Δt [s]", "Delta t [s]","Dt [s]"],
    "DO start [mg/L]": ["DO start [mg/L]"],
    "DO end [mg/L]": ["DO end [mg/L]"],
    "ΔDO [mg/L]": ["ΔDO [mg/L]", "Delta DO [mg/L]"],
    "slope [mg/L·s⁻¹]": ["slope [mg/L·s⁻¹]","slope [mg/L*s-1]","slope"],
    "OUR [mg/L·h]": ["OUR [mg/L·h]","OUR [mg/L*h]","OUR"],
    "R²": ["R²","R^2","R2"],
    "pH (średnie)": ["pH (średnie)","pH avg"],
    "T (średnia) [°C]": ["T (średnia) [°C]","T avg [°C]","T avg [C]","T avg"],
    "DO (min) [mg/L]": ["DO (min) [mg/L]","DO min [mg/L]"],
    "DO (max) [mg/L]": ["DO (max) [mg/L]","DO max [mg/L]"],
    "Liczba punktów": ["Liczba punktów","N punktów","N_punktow"],
    "SOUR_plateau": ["SOUR_plateau","SOUR_plateau_segment [gO2/kg sm·h]","SOUR plateau [gO2/kg sm·h]"],
    "AUC_segment": ["AUC_segment","AUC_segment [gO2/kg sm]","AUC [gO2/kg sm]"],
    "RMSE_segment": ["RMSE_segment","RMSE"],
    "%Inhibicji_vs_ref": ["%Inhibicji_vs_ref","% inhibicji segment [%]","% inhibicji [%]"],
    "N_punktow_segmentu": ["N_punktow_segmentu","N punktow segmentu"],
}

META_ALIAS = {
    "Data testu": ["Data testu","Data"],
    "Respirometr/wersja": ["Respirometr/wersja","Respirometr"],
    "Dawka Cr [mg]": ["Dawka Cr [mg]","Dawka"],
    "Ref_median_SOUR": ["Ref_median_SOUR","SOUR_ref_median"],
    "SOUR_plateau_avg": ["SOUR_plateau_avg","SOUR_plateau średnia","SOUR_plateau avg"],
    "% inhibicji test [%]": ["% inhibicji test [%]","%Inhibicji_vs_ref (test)","% inhibicji","%Inhibicji_vs_ref_avg"],
    "AUC_test [gO2/kg sm]": ["AUC_test [gO2/kg sm]","AUC test [gO2/kg sm]","AUC_test"],
    "Time_to_plateau [s]": ["Time_to_plateau [s]","Czas do plateau [s]"],
    "Arkusz meta": ["Arkusz meta"],
    "Arkusz surowy": ["Arkusz surowy"],
    "Ścieżka źródłowa": ["Ścieżka źródłowa","Sciezka zrodłowa","Ścieżka"],
    "Wersja skryptu": ["Wersja skryptu","Wersja"],
    "TestID": ["TestID","ID"],
    "Czas trwania [s]": ["Czas trwania [s]"],
    "Mediana kroku [s]": ["Mediana kroku [s]"],
    "Liczba punktów": ["Liczba punktów"],
    "DO min [mg/L] (całość)": ["DO min [mg/L] (całość)"],
    "DO max [mg/L] (całość)": ["DO max [mg/L] (całość)"],
    "DO avg [mg/L] (całość)": ["DO avg [mg/L] (całość)"],
    "T avg [°C] (całość)": ["T avg [°C] (całość)"],
    "pH avg (całość)": ["pH avg (całość)"],
    "ΔDO użyte [mg/L]": ["ΔDO użyte [mg/L]","Delta DO użyte [mg/L]"],
    "min R² użyte": ["min R² użyte","min R2 użyte"],
    "Segmenty zaakceptowane [#]": ["Segmenty zaakceptowane [#]","Acc segments [#]"],
    "Segmenty odrzucone przez R² [#]": ["Segmenty odrzucone przez R² [#]","Rejected R2 [#]"],
    "Segmenty odrzucone przez ΔDO [#]": ["Segmenty odrzucone przez ΔDO [#]","Rejected ΔDO [#]"],
}

def _norm(s:str)->str:
    return re.sub(r"\s+"," ", s.strip().lower()) if isinstance(s,str) else ""

def _to_float(val):
    if val is None or (isinstance(val,str) and not val.strip()):
        return np.nan
    if isinstance(val,(int,float,np.integer,np.floating)):
        return float(val)
    s=str(val).strip().replace(",", ".")
    s=re.sub(r"[^0-9\.\-eE]","",s)
    try: return float(s)
    except: return np.nan

def parse_time_to_seconds(x):
    if pd.isna(x): return np.nan
    if isinstance(x,(int,float,np.integer,np.floating)): return int(x)
    s = str(x).strip().replace(",",".")
    parts = s.split(":")
    try:
        if len(parts)==4:
            d,h,m,sec=parts; total=int(d)*86400+int(h)*3600+int(m)*60+float(sec)
        elif len(parts)==3:
            h,m,sec=parts; total=int(h)*3600+int(m)*60+float(sec)
        elif len(parts)==2:
            m,sec=parts; total=int(m)*60+float(sec)
        else:
            total=float(s)
        return int(round(total))
    except Exception:
        return np.nan

def _find_raw_sheet(xls: pd.ExcelFile)->Optional[str]:
    for s in xls.sheet_names:
        if _norm(s) in [_norm(c) for c in RAW_SHEET_CANDIDATES]:
            return s
    return None

def _load_clean_raw(meas_path: Path):
    xls = pd.ExcelFile(meas_path)
    raw_sheet = _find_raw_sheet(xls)
    if raw_sheet is None:
        return pd.DataFrame(), None
    df = pd.read_excel(meas_path, sheet_name=raw_sheet)
    cols = {c: str(c).strip() for c in df.columns}
    def find_col(cands):
        for name,norm in cols.items():
            for c in cands:
                if norm.lower()==c.lower(): return name
        for name,norm in cols.items():
            for c in cands:
                if c.lower() in norm.lower(): return name
        return None
    time_col = find_col(["Time (Day:Hour:Minute:Second)","Time"])
    do_col   = find_col(["DO (ppm)","DO"])
    t_col    = find_col(["T, (ºC)","T. (ºC)","T (°C)","T. (°C)","T. (oC)","T (oC)"])
    ph_col   = find_col(["pH"])
    if all([time_col,do_col,t_col,ph_col]):
        clean = df.iloc[1:, [df.columns.get_loc(time_col), df.columns.get_loc(do_col), df.columns.get_loc(t_col), df.columns.get_loc(ph_col)]].copy()
    else:
        n = len(df.columns)
        if n >= 5:
            clean = df.iloc[1:, [1,2,3,4]].copy()
        else:
            return pd.DataFrame(), raw_sheet
    clean.columns = ["Time (Day:Hour:Minute:Second)", "DO (ppm)", "T (°C)", "pH"]
    clean["Time (s)"] = clean["Time (Day:Hour:Minute:Second)"].apply(parse_time_to_seconds)
    for c in ["DO (ppm)","T (°C)","pH"]:
        clean[c]=pd.to_numeric(clean[c], errors="coerce")
    clean = clean.dropna(subset=["Time (s)","DO (ppm)"]).sort_values("Time (s)").reset_index(drop=True)
    return clean, raw_sheet

def _guess_meta_sheet(wb)->str:
    for s in wb.sheetnames:
        if "inhibicja" in _norm(s): return s
    for s in wb.sheetnames:
        if _norm(s) not in [_norm(x) for x in RAW_SHEET_CANDIDATES]:
            return s
    return wb.sheetnames[0]

def _find_label_value(sheet, labels, search_cols=20, search_rows=200):
    labs=[_norm(l) for l in labels]
    for r in range(1, search_rows+1):
        for c in range(1, search_cols+1):
            val = sheet.cell(row=r, column=c).value
            if _norm(val) in labs:
                for cc in range(c+1, c+5):
                    v2 = sheet.cell(row=r, column=cc).value
                    if v2 not in (None,""): return v2, (r,cc), f"label '{val}'"
                v2 = sheet.cell(row=r+1, column=c).value
                if v2 not in (None,""): return v2, (r+1,c), f"below '{val}'"
    return None, None, None

def _date_from_filename(stem:str):
    m = re.match(r"\s*(\d{1,2})\s+(\d{1,2})\s+(\d{4})", stem)
    if m:
        d,mn,y = m.groups()
        try: return pd.to_datetime(f"{y}-{mn}-{d}", errors="coerce")
        except: return pd.NaT
    return pd.NaT

def _valid_positive(x)->bool:
    f=_to_float(x); return (not pd.isna(f)) and (f>0)

def _load_meta_basic(meas_path: Path) -> Dict[str, object]:
    wb = load_workbook(meas_path, data_only=True)
    sname = _guess_meta_sheet(wb)
    sheet = wb[sname]
    meta = {"_sheet": sname, "_sources": {}}
    v = sheet["B2"].value
    date = pd.to_datetime(v, errors="coerce") if v not in (None,"") else pd.NaT
    if pd.isna(date):
        v_lab, rc, src = _find_label_value(sheet, ["data"])
        if v_lab not in (None,""):
            date = pd.to_datetime(v_lab, errors="coerce")
            if not pd.isna(date): meta["_sources"]["date"]=f"{sname}!{rc} via {src}"
    if pd.isna(date):
        date = _date_from_filename(meas_path.stem)
        if isinstance(date, pd.Timestamp) and not pd.isna(date):
            meta["_sources"]["date"]=f"filename:{meas_path.name}"
    if isinstance(date, pd.Timestamp) and not pd.isna(date):
        meta["date"]=date
    v = sheet["B3"].value
    if v not in (None,""): meta["respirometr"]=str(v).strip()
    v = sheet["C13"].value
    f=_to_float(v)
    if not pd.isna(f): meta["chrom_mg"]=f
    v59 = sheet["C59"].value; f59 = _to_float(v59)
    if (not pd.isna(f59)) and (f59>0):
        meta["dry_mass"]=f59
    else:
        v58 = sheet["C58"].value; v57 = sheet["C57"].value; vals=[]
        if _valid_positive(v58): vals.append(_to_float(v58))
        if _valid_positive(v57): vals.append(_to_float(v57))
        if len(vals)==2: meta["dry_mass"]=float(np.mean(vals))
        elif len(vals)==1: meta["dry_mass"]=float(vals[0])
    return meta

def _find_decreasing_runs(df: pd.DataFrame):
    t = df["Time (s)"].values; DO = df["DO (ppm)"].values
    n=len(df); allowed = np.ones(n, dtype=bool)
    for i in range(1,n):
        if DO[i]>DO[i-1] and t[i]>200:
            mask=(t>t[i])&(t<=t[i]+45); allowed[mask]=False
    runs=[]; idxs=np.where(allowed)[0]
    if len(idxs)==0: return runs
    rs=idxs[0]; prev=idxs[0]
    for cur in idxs[1:]:
        if DO[cur]>DO[prev] or cur!=prev+1:
            runs.append((rs,prev)); rs=cur
        prev=cur
    runs.append((rs,prev))
    return [(a,b) for a,b in runs if (b-a+1)>=3]

def _linreg(t,y):
    t=np.asarray(t,float); y=np.asarray(y,float)
    tm=t.mean(); ym=y.mean()
    den=((t-tm)**2).sum()
    if den==0: return {"slope":np.nan,"r2":np.nan,"rmse":np.nan,"n":0}
    b=((t-tm)*(y-ym)).sum()/den
    yfit = ym + b*(t-tm)
    ss_tot=((y-ym)**2).sum(); ss_res=((y-yfit)**2).sum()
    r2=1-ss_res/ss_tot if ss_tot>0 else np.nan
    rmse = math.sqrt(ss_res/len(t)) if len(t)>0 else np.nan
    return {"slope":b,"r2":r2,"rmse":rmse,"n":len(t)}

def _ph_avg_from_H(df_slice: pd.DataFrame)->float:
    H=(10**(-df_slice["pH"].astype(float))).mean()
    return -np.log10(H) if (H>0 and np.isfinite(H)) else np.nan

def analyze_segments_verbose(clean: pd.DataFrame, delta_do: float, min_r2: float):
    if clean is None or clean.empty:
        return pd.DataFrame(), 0, 0, 0
    runs=_find_decreasing_runs(clean)
    used=set(); segs=[]; t=clean["Time (s)"].values; do=clean["DO (ppm)"].values
    tol=0.001
    acc=0; rej_r2=0; rej_deltado=0
    for a,b in runs:
        i=a
        while i<b:
            if i in used: i+=1; continue
            j=i+1; chosen=None; hit_candidate=False
            while j<=b:
                if j in used: j+=1; continue
                d=do[i]-do[j]
                if (delta_do-tol)<=d<=(delta_do+tol):
                    hit_candidate=True
                    idx=list(range(i,j+1))
                    if any(k in used for k in idx): j+=1; continue
                    if any(do[k]>do[k-1] for k in range(i+1,j+1)): j+=1; continue
                    if len(idx)>=3:
                        st=_linreg(t[idx],do[idx])
                        if np.isfinite(st["r2"]) and st["r2"]>=min_r2:
                            chosen=(i,j,st); break
                        else:
                            rej_r2 += 1
                    j+=1; continue
                elif d<(delta_do-tol):
                    j+=1; continue
                else:
                    break
            if chosen:
                i0,j0,st=chosen
                part=clean.iloc[i0:j0+1]
                segs.append({
                    "t_start [s]": int(t[i0]),
                    "t_end [s]": int(t[j0]),
                    "Δt [s]": int(t[j0]-t[i0]),
                    "DO start [mg/L]": float(do[i0]),
                    "DO end [mg/L]": float(do[j0]),
                    "ΔDO [mg/L]": float(do[i0]-do[j0]),
                    "slope [mg/L·s⁻¹]": float(st["slope"]),
                    "OUR [mg/L·h]": float(-st["slope"]*3600.0),
                    "R²": float(st["r2"]),
                    "RMSE_segment": float(st["rmse"]),
                    "N_punktow_segmentu": int(st["n"]),
                    "pH (średnie)": float(_ph_avg_from_H(part)),
                    "T (średnia) [°C]": float(part["T (°C)"].mean()),
                    "DO (min) [mg/L]": float(part["DO (ppm)"].min()),
                    "DO (max) [mg/L]": float(part["DO (ppm)"].max()),
                    "Liczba punktów": int(len(part)),
                })
                for k in range(i0,j0+1): used.add(k)
                i=j0+1
                acc += 1
            else:
                if not hit_candidate:
                    rej_deltado += 1
                i+=1
    return pd.DataFrame(segs), acc, rej_r2, rej_deltado

def _read_block_BC(ws, r1, r2)->Dict[str, float]:
    out={}
    for r in range(r1, r2+1):
        lab = ws.cell(row=r, column=2).value
        val = ws.cell(row=r, column=3).value
        if isinstance(lab, str) and lab.strip():
            out[lab.strip()] = _to_float(val)
    return out

def _read_sm_direct_or_fallback(ws, c_row:int, d_row:int, block:Tuple[int,int])->Tuple[float,float]:
    out_total = _to_float(ws.cell(row=c_row, column=3).value)
    out_organic = _to_float(ws.cell(row=d_row, column=4).value)
    total = out_total if (not pd.isna(out_total) and out_total>0) else np.nan
    organic = out_organic if (not pd.isna(out_organic) and out_organic>0) else np.nan
    if not (not pd.isna(total) and total>0) or not (not pd.isna(organic) and organic>0):
        r1,r2 = block
        vals_C=[]; vals_D=[]
        for r in range(r1, r2+1):
            vC=_to_float(ws.cell(row=r, column=3).value)
            vD=_to_float(ws.cell(row=r, column=4).value)
            if not pd.isna(vC) and vC>0: vals_C.append(vC)
            if not pd.isna(vD) and vD>0: vals_D.append(vD)
        if (pd.isna(total) or total<=0) and len(vals_C)>0:
            total=float(np.mean(vals_C))
        if (pd.isna(organic) or organic<=0) and len(vals_D)>0:
            organic=float(np.mean(vals_D))
    return total, organic

def _build_meta_row(meas_path: Path, clean: pd.DataFrame, raw_sheet_name: Optional[str], delta_do: float, min_r2: float, acc: int, rej_r2: int, rej_deltado: int) -> Dict[str, object]:
    wb = load_workbook(meas_path, data_only=True)
    sname = _guess_meta_sheet(wb)
    ws = wb[sname]
    meta_basic = _load_meta_basic(meas_path)
    row = {
        "Data testu": (pd.to_datetime(meta_basic.get("date")).date()
                       if isinstance(meta_basic.get("date"), pd.Timestamp) and not pd.isna(meta_basic.get("date")) else None),
        "Respirometr/wersja": meta_basic.get("respirometr"),
        "Dawka Cr [mg]": _to_float(meta_basic.get("chrom_mg")),
    }
    for k,v in _read_block_BC(ws, 6, 38).items(): row[k]=v
    for k,v in _read_block_BC(ws, 41, 44).items(): row[f"{k} (respiro)"]=v
    for k,v in _read_block_BC(ws, 47, 53).items(): row[f"{k} (mech. ocz.)"]=v
    sm, smo = _read_sm_direct_or_fallback(ws, c_row=59, d_row=59, block=(56,59))
    if not pd.isna(sm):  row["średnia sucha masa (respiro)"] = sm
    if not pd.isna(smo): row["średnia sucha masa organiczna (respiro)"] = smo
    sm, smo = _read_sm_direct_or_fallback(ws, c_row=66, d_row=66, block=(63,66))
    if not pd.isna(sm):  row["średnia sucha masa (osad)"] = sm
    if not pd.isna(smo): row["średnia sucha masa organiczna (osad)"] = smo
    sm, smo = _read_sm_direct_or_fallback(ws, c_row=72, d_row=72, block=(69,72))
    if not pd.isna(sm):  row["średnia sucha masa (placek)"] = sm
    if not pd.isna(smo): row["średnia sucha masa organiczna (placek)"] = smo

    if clean is not None and (not clean.empty):
        tt = clean["Time (s)"].values
        dd = clean["DO (ppm)"].values
        row["Czas trwania [s]"] = int(np.nanmax(tt)-np.nanmin(tt)) if len(tt)>1 else np.nan
        diffs = np.diff(np.sort(tt))
        row["Mediana kroku [s]"] = float(np.median(diffs)) if len(diffs)>0 else np.nan
        row["Liczba punktów"] = int(len(clean))
        row["DO min [mg/L] (całość)"] = float(np.nanmin(dd)) if len(dd)>0 else np.nan
        row["DO max [mg/L] (całość)"] = float(np.nanmax(dd)) if len(dd)>0 else np.nan
        row["DO avg [mg/L] (całość)"] = float(np.nanmean(dd)) if len(dd)>0 else np.nan
        row["T avg [°C] (całość)"] = float(np.nanmean(clean["T (°C)"].values)) if "T (°C)" in clean.columns else np.nan
        try:
            H = (10**(-clean["pH"].astype(float))).mean()
            row["pH avg (całość)"] = float(-np.log10(H)) if H>0 else np.nan
        except Exception:
            row["pH avg (całość)"] = np.nan
    else:
        row["Czas trwania [s]"] = np.nan
        row["Mediana kroku [s]"] = np.nan
        row["Liczba punktów"] = np.nan
        row["DO min [mg/L] (całość)"] = np.nan
        row["DO max [mg/L] (całość)"] = np.nan
        row["DO avg [mg/L] (całość)"] = np.nan
        row["T avg [°C] (całość)"] = np.nan
        row["pH avg (całość)"] = np.nan

    row["ΔDO użyte [mg/L]"] = float(delta_do)
    row["min R² użyte"] = float(min_r2)
    row["Segmenty zaakceptowane [#]"] = int(acc)
    row["Segmenty odrzucone przez R² [#]"] = int(rej_r2)
    row["Segmenty odrzucone przez ΔDO [#]"] = int(rej_deltado)

    row["Arkusz meta"] = ws.title
    row["Arkusz surowy"] = raw_sheet_name
    row["Ścieżka źródłowa"] = str(meas_path)
    row["Wersja skryptu"] = VERSION
    row["TestID"] = meas_path.stem
    return row

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
def _q(tag): return f"{{{NS_MAIN}}}{tag}"

def _last_used_row_any(ws, header_row:int=1)->int:
    for r in range(ws.max_row, header_row, -1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        if any(v not in (None,"") for v in row_vals):
            return r
    return header_row

def _last_used_col_in_header(ws, header_row:int=1)->int:
    for c in range(ws.max_column, 0, -1):
        v = ws.cell(row=header_row, column=c).value
        if v not in (None,""):
            return c
    return 1

def _rewrite_table_xml_with_headers(xlsx_path: Path, table_name: str, new_ref: str, headers: List[str])->bool:
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            match_name=None; xml_old=None
            table_xmls=[n for n in zin.namelist() if n.startswith("xl/tables/table") and n.endswith(".xml")]
            for tn in table_xmls:
                data=zin.read(tn)
                root=ET.fromstring(data)
                dn = root.attrib.get("displayName") or root.attrib.get(f"{{{NS_MAIN}}}displayName")
                if dn == table_name:
                    match_name=tn; xml_old=data; break
            if not match_name:
                return False
        root=ET.fromstring(xml_old)
        root.attrib["ref"] = new_ref
        root.attrib["headerRowCount"] = "1"
        tcols = root.find(_q("tableColumns"))
        if tcols is None:
            tcols = ET.SubElement(root, _q("tableColumns"))
        for ch in list(tcols): tcols.remove(ch)
        tcols.attrib["count"]=str(len(headers))
        for i,h in enumerate(headers, start=1):
            col = ET.SubElement(tcols, _q("tableColumn"))
            col.attrib["id"]=str(i)
            col.attrib["name"]=str(h if h is not None else f"Kol_{i}")
        new_xml = ET.tostring(root, encoding="utf-8", xml_exclamation=True) if False else ET.tostring(root, encoding="utf-8", xml_declaration=True)
        tmp = xlsx_path.with_suffix(".tmp.xlsx")
        with zipfile.ZipFile(xlsx_path,'r') as zin, zipfile.ZipFile(tmp,'w') as zout:
            for item in zin.infolist():
                if item.filename == match_name:
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))
        tmp.replace(xlsx_path)
        return True
    except Exception:
        return False

def _expand_table_sync_columns(xlsx_path: Path, sheet_name: str, table_name: str, header_row:int=1)->bool:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames: return False
    ws = wb[sheet_name]
    last_row = _last_used_row_any(ws, header_row=header_row)
    last_col = _last_used_col_in_header(ws, header_row=header_row)
    new_ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, last_col+1)]
    ok = _rewrite_table_xml_with_headers(xlsx_path, table_name, new_ref, headers)
    return ok

def _build_col_map(ws, alias_map: Dict[str, List[str]])->Dict[str,int]:
    header_row=1
    headers=[ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column+1)]
    headers=[h if h is not None else "" for h in headers]
    nh=[_norm(h) for h in headers]
    mapping={}
    for canon, aliases in alias_map.items():
        found=None
        for a in aliases:
            an=_norm(a)
            for i,hh in enumerate(nh):
                if hh==an:
                    found=i+1; break
            if found: break
        if not found:
            for a in aliases:
                an=_norm(a)
                for i,hh in enumerate(nh):
                    if an and an in hh:
                        found=i+1; break
                if found: break
        if found:
            mapping[canon]=found
    return mapping

def append_rows_to_db(edit_path: Path, rows: List[Dict]):
    wb = load_workbook(edit_path)
    if "Dane" not in wb.sheetnames:
        raise RuntimeError("W bazie nie znaleziono arkusza 'Dane'.")
    ws = wb["Dane"]
    header_row=1
    col_map = _build_col_map(ws, SEGMENT_ALIAS)

    last_before = _last_used_row_any(ws, header_row=header_row)
    row_ptr = last_before; added=0
    for row in rows:
        row_ptr += 1
        wrote=False
        for canon_key, col_index in col_map.items():
            if canon_key in row:
                val=row.get(canon_key)
                cell = ws.cell(row=row_ptr, column=col_index, value=val)
                if canon_key=="Data" and val not in (None,"") and hasattr(cell,"number_format"):
                    try: cell.number_format = DATE_EXCEL_FMT
                    except Exception: pass
                wrote=True
        if wrote: added += 1
        else: row_ptr -= 1
    if added>0:
        wb.save(edit_path)
    return added

def _ensure_sheet(wb, name:str):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)

def _append_meta_row_any_header(edit_path: Path, meta_row: Dict[str, object], sheet_name:str="Testy (meta)"):
    wb = load_workbook(edit_path)
    ws = _ensure_sheet(wb, sheet_name)
    header_row=1
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column+1)]
    headers = [h if h is not None else "" for h in headers]

    last = 1
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(row=r, column=c).value not in (None,"") for c in range(1, ws.max_column+1)):
            last=r; break
    target = max(last+1, 2)

    wrote=False
    alias_to_canon={}
    for canon, aliases in META_ALIAS.items():
        for a in aliases:
            alias_to_canon[_norm(a)] = canon

    for c, H in enumerate(headers, start=1):
        v=None
        if H in meta_row:
            v=meta_row[H]
        else:
            canon = alias_to_canon.get(_norm(H))
            if canon and (canon in meta_row):
                v=meta_row[canon]
        if v is not None:
            cell = ws.cell(row=target, column=c, value=v)
            if str(H).strip().lower().startswith("data"):
                try: cell.number_format = DATE_EXCEL_FMT
                except Exception: pass
            wrote=True

    if wrote:
        wb.save(edit_path)
    else:
        wb.save(edit_path)

def _propagate_ST_formulas(edit_path: Path, sheet_name:str="Dane"):
    wb = load_workbook(edit_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close(); return
    ws = wb[sheet_name]
    header_row=1
    last_row = _last_used_row_any(ws, header_row=header_row)
    if last_row <= 2:
        wb.close(); return
    s2 = ws.cell(row=2, column=19).value
    t2 = ws.cell(row=2, column=20).value
    for r in range(3, last_row+1):
        if isinstance(s2, str) and s2.startswith("="):
            ws.cell(row=r, column=19).value = s2
        if isinstance(t2, str) and t2.startswith("="):
            ws.cell(row=r, column=20).value = t2
    wb.save(edit_path)
    wb.close()

def process_file(meas_path: Path, delta_do: float, min_r2: float):
    clean, raw_sheet = _load_clean_raw(meas_path)
    seg_df, acc, rej_r2, rej_deltado = analyze_segments_verbose(clean, delta_do=delta_do, min_r2=min_r2)
    meta_basic=_load_meta_basic(meas_path)

    rows=[]
    for _,r in seg_df.iterrows():
        rows.append({
            "Data": (pd.to_datetime(meta_basic.get("date")).date() if isinstance(meta_basic.get("date"), pd.Timestamp) and not pd.isna(meta_basic.get("date")) else None),
            "Dawka Cr [mg]": _to_float(meta_basic.get("chrom_mg")),
            "Respirometr/wersja": meta_basic.get("respirometr"),
            "sm (średnia) g sm/L": _to_float(meta_basic.get("dry_mass")),
            "t_start [s]": int(r["t_start [s]"]),
            "t_end [s]": int(r["t_end [s]"]),
            "Δt [s]": int(r["Δt [s]"]),
            "DO start [mg/L]": float(r["DO start [mg/L]"]),
            "DO end [mg/L]": float(r["DO end [mg/L]"]),
            "ΔDO [mg/L]": float(r["ΔDO [mg/L]"]),
            "slope [mg/L·s⁻¹]": float(r["slope [mg/L·s⁻¹]"]),
            "OUR [mg/L·h]": float(-r["slope [mg/L·s⁻¹]"]*3600.0),
            "R²": float(r["R²"]),
            "pH (średnie)": float(r["pH (średnie)"]),
            "T (średnia) [°C]": float(r["T (średnia) [°C]"]),
            "DO (min) [mg/L]": float(r["DO (min) [mg/L]"]),
            "DO (max) [mg/L]": float(r["DO (max) [mg/L]"]),
            "Liczba punktów": int(r["Liczba punktów"]),
            "N_punktow_segmentu": int(r["N_punktow_segmentu"]),
            "RMSE_segment": float(r["RMSE_segment"]),
        })
    sm = _to_float(meta_basic.get("dry_mass"))
    dose = _to_float(meta_basic.get("chrom_mg"))
    for rr in rows:
        OUR = _to_float(rr.get("OUR [mg/L·h]"))
        if (not pd.isna(sm)) and sm>0 and (not pd.isna(OUR)):
            sour = OUR / sm
            rr["SOUR_plateau"] = sour
            dt = _to_float(rr.get("Δt [s]"))
            rr["AUC_segment"] = sour * (dt/3600.0) if (not pd.isna(dt)) else np.nan
        else:
            rr["SOUR_plateau"] = np.nan
            rr["AUC_segment"] = np.nan
        if (not pd.isna(dose)) and abs(dose) < 1e-12:
            rr["%Inhibicji_vs_ref"] = np.nan
        else:
            rr["%Inhibicji_vs_ref"] = np.nan

    meta_row = _build_meta_row(meas_path, clean, raw_sheet, delta_do, min_r2, acc, rej_r2, rej_deltado)
    return rows, meta_basic, meta_row, raw_sheet, clean, (acc, rej_r2, rej_deltado)

def main():
    ap = argparse.ArgumentParser(description="Import segmentów + meta; cache JSON; aliasy nagłówków; propagacja S/T; log problemów; bez % inhibicji dla referencji.")
    ap.add_argument("--input-dir", required=True)
    ap.add_argument("--db", required=True)
    ap.add_argument("--delta-do", type=float, default=0.25)
    ap.add_argument("--min-r2", type=float, default=0.95)
    ap.add_argument("--out-copy", default=None)
    ap.add_argument("--table-name-hint", default="Data")
    ap.add_argument("--meta-sheet", default="Testy (meta)")
    ap.add_argument("--meta-table-name", default="Testy_meta")
    ap.add_argument("--cache-json", default=None, help="Ścieżka do JSON z listą już przetworzonych plików (domyślnie: 'processed_files.json' w --input-dir)")
    args = ap.parse_args()

    input_dir=Path(args.input-dir) if hasattr(args,'input-dir') else Path(args.input_dir)
    cache_path = Path(args.cache_json) if args.cache_json else (input_dir / "processed_files.json")
    db_path=Path(args.db); out_copy=Path(args.out_copy) if args.out_copy else None
    if not input_dir.exists(): print(f"Błąd: folder {input_dir} nie istnieje.", file=sys.stderr); sys.exit(1)
    if not db_path.exists(): print(f"Błąd: plik bazy {db_path} nie istnieje.", file=sys.stderr); sys.exit(1)

    edit_path = out_copy if out_copy and out_copy.exists() else (out_copy if out_copy else db_path)
    if out_copy and not out_copy.exists():
        out_copy.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(db_path, out_copy)

    files=[p for p in input_dir.glob("*.xlsx") if p.name.lower()!=db_path.name.lower() and (out_copy is None or p.name.lower()!=edit_path.name.lower()) and "_analiza" not in p.stem.lower()]
    if not files:
        print("Nie znaleziono plików .xlsx do analizy w podanym folderze."); sys.exit(0)

    # Wczytaj cache
    try:
        if cache_path.exists():
            processed_cache = json.loads(cache_path.read_text(encoding="utf-8"))
        else:
            processed_cache = {}
    except Exception:
        processed_cache = {}

    all_segment_rows = []
    all_meta_rows = []

    missing_sm = set()
    no_segments = set()
    total_meta_appended = 0

    for f in sorted(files):
        sig = {"size": f.stat().st_size, "mtime": int(f.stat().st_mtime)}
        cache_key = f.name
        if cache_key in processed_cache:
            prev = processed_cache[cache_key]
            if prev.get("size")==sig["size"] and prev.get("mtime")==sig["mtime"]:
                print(f"[{f.name}] POMINIĘTO (bez zmian według cache).")
                continue

        try:
            rows, meta_basic, meta_row, raw_sname, clean, counts = process_file(f, delta_do=args.delta_do, min_r2=args.min_r2)
            acc, rej_r2, rej_deltado = counts
            if acc==0:
                no_segments.add(f.name)
            if _to_float(meta_basic.get("dry_mass"))<=0 or pd.isna(_to_float(meta_basic.get("dry_mass"))):
                missing_sm.add(f.name)
            all_segment_rows.append((f.name, meta_basic, rows))
            all_meta_rows.append((f.name, meta_row))
            print(f"[{f.name}] segm={len(rows)} (acc={acc}, rej_R2={rej_r2}, rej_ΔDO={rej_deltado}); meta=OK")
            processed_cache[cache_key]=sig
        except Exception as e:
            print(f"[{f.name}] BŁĄD: {e}", file=sys.stderr)

    # referencje
    ref_median = {}
    ref_vals = {}
    for (fname, meta_basic, rows) in all_segment_rows:
        d = (pd.to_datetime(meta_basic.get("date")).date() if isinstance(meta_basic.get("date"), pd.Timestamp) and not pd.isna(meta_basic.get("date")) else None)
        resp = meta_basic.get("respirometr")
        dose = _to_float(meta_basic.get("chrom_mg"))
        if (dose is not None) and (not pd.isna(dose)) and abs(dose) < 1e-12:
            key=(d,resp)
            for r in rows:
                v=_to_float(r.get("SOUR_plateau"))
                if not pd.isna(v):
                    ref_vals.setdefault(key, []).append(v)
    for k,vals in ref_vals.items():
        if len(vals)>0: ref_median[k]=float(np.median(vals))

    # % inhibicji (segmenty, tylko nie-ref)
    for (fname, meta_basic, rows) in all_segment_rows:
        d = (pd.to_datetime(meta_basic.get("date")).date() if isinstance(meta_basic.get("date"), pd.Timestamp) and not pd.isna(meta_basic.get("date")) else None)
        resp = meta_basic.get("respirometr")
        dose = _to_float(meta_basic.get("chrom_mg"))
        if (dose is not None) and (not pd.isna(dose)) and abs(dose) < 1e-12:
            continue
        ref = ref_median.get((d,resp), np.nan)
        if pd.isna(ref) or ref<=0:
            continue
        for r in rows:
            sourp = _to_float(r.get("SOUR_plateau"))
            if not pd.isna(sourp):
                r["%Inhibicji_vs_ref"] = 100.0*(1.0 - sourp / ref)

    # test-level
    test_summary = {}
    for (fname, meta_basic, rows) in all_segment_rows:
        d = (pd.to_datetime(meta_basic.get("date")).date() if isinstance(meta_basic.get("date"), pd.Timestamp) and not pd.isna(meta_basic.get("date")) else None)
        resp = meta_basic.get("respirometr")
        dose = _to_float(meta_basic.get("chrom_mg"))
        key_test=(d, resp, dose, Path(fname).stem)
        ref = ref_median.get((d,resp), np.nan)
        starts = [ _to_float(r.get("t_start [s]")) for r in rows if not pd.isna(_to_float(r.get("t_start [s]"))) ]
        series_start = float(np.min(starts)) if len(starts)>0 else np.nan
        first_pl = np.nan
        for r in rows:
            sourp = _to_float(r.get("SOUR_plateau"))
            if pd.isna(first_pl) and (not pd.isna(sourp)):
                first_pl = _to_float(r.get("t_start [s]"))
        sum_auc = float(np.nansum([_to_float(r.get("AUC_segment")) for r in rows])) if len(rows)>0 else np.nan
        lst_sour = [ _to_float(r.get("SOUR_plateau")) for r in rows ]
        lst_sour = [v for v in lst_sour if not pd.isna(v)]
        test_summary[key_test] = {
            "Ref_median_SOUR": ref if (not pd.isna(ref)) else np.nan,
            "SOUR_plateau_avg": float(np.mean(lst_sour)) if len(lst_sour)>0 else np.nan,
            "AUC_test [gO2/kg sm]": sum_auc,
            "Time_to_plateau [s]": (first_pl - series_start) if (not pd.isna(first_pl) and not pd.isna(series_start)) else np.nan,
        }

    # zapis segmentów
    edit_path = out_copy if (out_copy and out_copy.exists()) else (out_copy if out_copy else db_path)
    total_added=0
    for (fname, meta_basic, rows) in all_segment_rows:
        if rows:
            total_added += append_rows_to_db(edit_path, rows)

    # zapis meta + % inhibicji test (bez referencji)
    total_meta_appended=0
    for (fname, meta_row) in all_meta_rows:
        d = meta_row.get("Data testu", None)
        resp = meta_row.get("Respirometr/wersja", None)
        dose = _to_float(meta_row.get("Dawka Cr [mg]"))
        key_test=(d, resp, dose, Path(meta_row.get("TestID", Path(fname).stem)).stem)
        add = test_summary.get(key_test, {})
        sp = add.get("SOUR_plateau_avg", np.nan)
        rf = add.get("Ref_median_SOUR", np.nan)
        if (dose is not None) and (not pd.isna(dose)) and abs(dose) < 1e-12:
            inh = np.nan
        else:
            inh = 100.0*(1.0 - sp/rf) if (not pd.isna(sp) and not pd.isna(rf) and rf>0) else np.nan
        meta_row_out = dict(meta_row)
        meta_row_out.update({
            "Ref_median_SOUR": add.get("Ref_median_SOUR", np.nan),
            "SOUR_plateau_avg": add.get("SOUR_plateau_avg", np.nan),
            "% inhibicji test [%]": inh,
            "AUC_test [gO2/kg sm]": add.get("AUC_test [gO2/kg sm]", np.nan),
            "Time_to_plateau [s]": add.get("Time_to_plateau [s]", np.nan),
        })
        _append_meta_row_any_header(edit_path, meta_row_out, sheet_name=args.meta_sheet)
        total_meta_appended += 1

    # Rozszerz zakres tabel
    try:
        _expand_table_sync_columns(edit_path, sheet_name="Dane", table_name=args.table_name_hint, header_row=1)
    except Exception:
        pass
    try:
        _expand_table_sync_columns(edit_path, sheet_name=args.meta_sheet, table_name=args.meta_table_name, header_row=1)
    except Exception:
        pass

    # Propagacja formuł S/T w "Dane"
    try:
        _propagate_ST_formulas(edit_path, sheet_name="Dane")
    except Exception as e:
        print(f"Uwaga: nie udało się przepisać formuł S/T: {e}", file=sys.stderr)

    # problems_log.txt
    plog = Path("problems_log.txt")
    lines=[]
    if missing_sm:
        lines.append("Brak suchej masy (SM) – nie policzono SOUR/AUC dla plików:")
        for n in sorted(missing_sm): lines.append(f"  - {n}")
        lines.append("")
    if no_segments:
        lines.append("Brak zaakceptowanych segmentów (ΔDO / R²):")
        for n in sorted(no_segments): lines.append(f"  - {n}")
        lines.append("")
    if not lines:
        lines = ["(brak problemów)"]
    plog.write_text("\n".join(lines), encoding="utf-8")

    # Podsumowanie
    print("\n=== Podsumowanie problemów ===")
    if missing_sm: print(f" Brak SM: {len(missing_sm)} plik(ów)")
    if no_segments: print(f" Brak zaakceptowanych segmentów: {len(no_segments)} plik(ów)")
    if (not missing_sm) and (not no_segments): print(" (brak)")
    print(f"(Zapisano listę problemów do: {plog})")
    print(f"\nZakończono. Segmenty dopisane: {total_added}; metadane dopisane: {total_meta_appended} wierszy.\nZapisano do: {edit_path}")

    # Zapisz/aktualizuj cache
    try:
        cache_path.write_text(json.dumps(processed_cache, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Zapisano cache: {cache_path}")
    except Exception as e:
        print(f"Uwaga: nie udało się zapisać cache ({cache_path}): {e}", file=sys.stderr)

if __name__ == "__main__":
    main()
